import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook

from meal_planner.nutrition_catalog import NUTRIENT_HEADER_BY_KEY
from meal_planner.nutrition_db import (
    database_path,
    load_catalog_entries,
    load_nutrition_profile,
    load_target_settings,
    load_target_rows,
    save_catalog_entries,
    save_nutrition_profile,
    save_target_settings,
    save_target_rows,
)
from meal_planner.settings import clear_settings_cache, get_settings


def _write_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        ws.cell(row, col).value = value


def _make_workbook() -> Workbook:
    settings = get_settings()
    wb = Workbook()
    menu_ws = wb.active
    menu_ws.title = settings.sheets.menu_v5
    nutrition_ws = wb.create_sheet(settings.sheets.nutrition_list)

    nutrient_headers = list(NUTRIENT_HEADER_BY_KEY.values())
    nutrient_values = list(range(100, 100 + len(nutrient_headers)))
    _write_row(nutrition_ws, 1, ["類別", "名稱", "暫停", "Min (g)", "Max (g)", "DayMax (g)", *nutrient_headers])
    _write_row(nutrition_ws, 2, ["水果", "蘋果", "", 50, 200, 300, *nutrient_values])

    first_col = settings.menu_v5_layout.nutrient_first_col
    for idx, header in enumerate(nutrient_headers):
        menu_ws.cell(settings.menu_v5_layout.indicator_header_row, first_col + idx).value = header
        menu_ws.cell(settings.menu_v5_layout.workday_indicator_row, first_col + idx).value = f"W{idx}"
        menu_ws.cell(settings.menu_v5_layout.nonworkday_indicator_row, first_col + idx).value = f"N{idx}"
    return wb


class NutritionDatabaseTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_root = os.environ.get("MENU_PROJECT_ROOT")
        os.environ["MENU_PROJECT_ROOT"] = self.tmp.name
        clear_settings_cache()

    def tearDown(self):
        if self.old_root is None:
            os.environ.pop("MENU_PROJECT_ROOT", None)
        else:
            os.environ["MENU_PROJECT_ROOT"] = self.old_root
        clear_settings_cache()
        self.tmp.cleanup()

    def test_catalog_bootstraps_once_from_workbook(self):
        settings = get_settings()
        wb = _make_workbook()

        first = load_catalog_entries(settings, wb)
        wb[settings.sheets.nutrition_list].cell(2, 2).value = "改名"
        second = load_catalog_entries(settings, wb)

        self.assertTrue(database_path(settings).is_file())
        self.assertEqual(first[0].name, "蘋果")
        self.assertEqual(second[0].name, "蘋果")
        self.assertEqual(second[0].nutrients["kcal"], 100.0)

    def test_save_catalog_replaces_rows_and_assigns_new_row_index(self):
        settings = get_settings()
        wb = _make_workbook()
        original = load_catalog_entries(settings, wb)[0]

        saved = save_catalog_entries(
            [
                {
                    "row_index": original.row_index,
                    "paused": True,
                    "category": "水果",
                    "name": "青蘋果",
                    "min_g": "40",
                    "max_g": "180",
                    "daymax_g": "",
                    "nutrients": {key: original.nutrients[key] for key in original.nutrients},
                },
                {
                    "category": "飲品",
                    "name": "豆奶",
                    "min_g": "",
                    "max_g": "",
                    "daymax_g": "300",
                    "nutrients": {"kcal": "50", "protein_g": "4"},
                },
            ],
            settings,
        )

        self.assertEqual([entry.name for entry in saved], ["青蘋果", "豆奶"])
        self.assertTrue(saved[0].paused)
        self.assertEqual(saved[0].min_g, 40.0)
        self.assertIsNone(saved[0].daymax_g)
        self.assertEqual(saved[1].row_index, 3)
        self.assertEqual(saved[1].nutrients["kcal"], 50.0)
        self.assertEqual(saved[1].nutrients["fat_trans_g"], 0.0)
        with self.assertRaises(ValueError):
            save_catalog_entries([{"category": "Missing name"}], settings)

    def test_targets_bootstrap_once_from_workbook(self):
        settings = get_settings()
        wb = _make_workbook()
        first_col = settings.menu_v5_layout.nutrient_first_col

        first = load_target_rows(settings, wb)
        wb[settings.sheets.menu_v5].cell(settings.menu_v5_layout.workday_indicator_row, first_col).value = "changed"
        second = load_target_rows(settings, wb)

        self.assertEqual(first[0][0], NUTRIENT_HEADER_BY_KEY["kcal"])
        self.assertEqual(first[1][0], "W0")
        self.assertEqual(first[2][0], "N0")
        self.assertEqual(second[1][0], "W0")

    def test_save_targets_replaces_sqlite_rows_and_validates_indicator_text(self):
        settings = get_settings()
        wb = _make_workbook()
        headers, _, _ = load_target_rows(settings, wb)
        workday = ["100-200", "10-20", "20-30", "< 5", "< 10", "< 100", "> 300", "< 27.5% kcal", "< 7% kcal", "< 1% kcal"]
        nonworkday = ["90-180", "9-18", "18-28", "< 4", "< 9", "< 90", "> 280", "< 27.5% kcal", "< 7% kcal", "< 1% kcal"]

        _, saved_workday, saved_nonworkday = save_target_rows(headers, workday, nonworkday, settings)

        self.assertEqual(saved_workday, workday)
        self.assertEqual(saved_nonworkday, nonworkday)
        with self.assertRaises(ValueError):
            save_target_rows(headers, [""] + workday[1:], nonworkday, settings)

    def test_target_settings_round_trips_through_sqlite(self):
        settings = get_settings()

        before = load_target_settings(settings)
        saved = save_target_settings(
            {
                "workday": {"activity_factor": 1.42, "calorie_range_band": 60, "sodium_mg": 1900},
                "nonworkday": {"activity_factor": 1.18, "sugar_g": 45, "fat_sat_pct": 8},
            },
            settings,
        )
        after = load_target_settings(settings)

        self.assertEqual(before["workday"]["activity_factor"], 1.35)
        self.assertEqual(saved["workday"]["activity_factor"], 1.42)
        self.assertEqual(saved["workday"]["calorie_range_band"], 60)
        self.assertEqual(saved["workday"]["sodium_mg"], 1900)
        self.assertEqual(saved["nonworkday"]["activity_factor"], 1.18)
        self.assertEqual(saved["nonworkday"]["sugar_g"], 45)
        self.assertEqual(saved["nonworkday"]["fat_sat_pct"], 8)
        self.assertEqual(after, saved)
        with self.assertRaises(ValueError):
            save_target_settings({"workday": {"activity_factor": -1}}, settings)

    def test_nutrition_profile_round_trips_through_sqlite(self):
        settings = get_settings()
        today = datetime.now(ZoneInfo("Asia/Hong_Kong")).date()
        dob_42 = f"{today.year - 42:04d}-01-01"
        dob_43 = f"{today.year - 43:04d}-01-01"

        before = load_nutrition_profile(settings)
        saved = save_nutrition_profile(
            {
                "dob": dob_42,
                "gender": "female",
                "height_cm": 165.5,
                "weight_history": [{"weight_kg": 58.2, "recorded_at": "2026-06-20 10:00:00"}],
            },
            settings,
        )
        same_weight = save_nutrition_profile(
            {
                "dob": dob_42,
                "gender": "female",
                "height_cm": 165.5,
                "weight_history": [{"weight_kg": 58.2, "recorded_at": "2026-06-20 10:00:00"}],
            },
            settings,
        )
        age_changed = save_nutrition_profile(
            {
                "dob": dob_43,
                "gender": "female",
                "height_cm": 165.5,
                "weight_history": [{"weight_kg": 58.2, "recorded_at": "2026-06-20 10:00:00"}],
            },
            settings,
        )
        changed_weight = save_nutrition_profile(
            {
                "dob": dob_43,
                "gender": "female",
                "height_cm": 165.5,
                "weight_history": [
                    {"weight_kg": 58.2, "recorded_at": "2026-06-20 10:00:00"},
                    {"weight_kg": 59.1, "recorded_at": "2026-06-21 10:00:00"},
                ],
            },
            settings,
        )
        after = load_nutrition_profile(settings)

        self.assertEqual(
            before,
            {"age": None, "dob": "", "gender": "", "height_cm": None, "weight_kg": None, "last_updated": "", "weight_history": []},
        )
        self.assertEqual(saved["dob"], dob_42)
        self.assertEqual(saved["age"], 42)
        self.assertEqual(saved["gender"], "female")
        self.assertEqual(saved["height_cm"], 165.5)
        self.assertEqual(saved["weight_kg"], 58.2)
        self.assertEqual(saved["last_updated"], "2026-06-20 10:00:00")
        self.assertEqual([item["weight_kg"] for item in saved["weight_history"]], [58.2])
        self.assertEqual([item["weight_kg"] for item in same_weight["weight_history"]], [58.2])
        self.assertEqual(age_changed["dob"], dob_43)
        self.assertEqual(age_changed["age"], 43)
        self.assertEqual([item["weight_kg"] for item in age_changed["weight_history"]], [58.2])
        self.assertEqual(changed_weight["weight_kg"], 59.1)
        self.assertEqual(changed_weight["last_updated"], "2026-06-21 10:00:00")
        self.assertEqual([item["weight_kg"] for item in changed_weight["weight_history"]], [58.2, 59.1])
        self.assertEqual(after, changed_weight)
        with self.assertRaises(ValueError):
            save_nutrition_profile({"gender": "unknown"}, settings)


if __name__ == "__main__":
    unittest.main()
