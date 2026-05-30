import os
import tempfile
import unittest

from openpyxl import Workbook

from meal_planner.excel_io import load_roster_map
from meal_planner.maintenance_db import (
    bootstrap_roster_code_definitions,
    import_runtime_inputs_from_workbook,
    list_runtime_input_status,
    load_roster_code_definitions,
    load_sheet_rows,
    save_roster_code_definitions,
    save_sheet_rows,
)
from meal_planner.settings import clear_settings_cache, get_settings


class MaintenanceDatabaseTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_root = os.environ.get("MENU_PROJECT_ROOT")
        os.environ["MENU_PROJECT_ROOT"] = self.tmp.name
        clear_settings_cache()

    def tearDown(self):
        if self.old_root is None:
            os.environ.pop("MENU_PROJECT_ROOT", None)
        else:
            os.environ["MENU_PROJECT_ROOT"] = self.old_root
        clear_settings_cache()
        self.tmp.cleanup()

    def test_sheet_bootstraps_once_then_uses_sqlite_copy(self):
        settings = get_settings()
        wb = Workbook()
        wb.active.title = settings.sheets.public_holidays
        ws = wb[settings.sheets.public_holidays]
        ws.cell(1, 1).value = "日期"
        ws.cell(1, 2).value = "假期名稱"
        ws.cell(2, 1).value = "2026-01-01"
        ws.cell(2, 2).value = "元旦"

        first = load_sheet_rows("public_holidays", settings, wb)
        ws.cell(2, 2).value = "Changed"
        second = load_sheet_rows("public_holidays", settings, wb)

        self.assertEqual(first["rows"][1], ["2026-01-01", "元旦"])
        self.assertEqual(second["rows"][1], ["2026-01-01", "元旦"])

    def test_save_replaces_sheet_rows(self):
        settings = get_settings()

        result = save_sheet_rows("overtime", [["日期", "開工"], ["2026-05-23", "09:00"]], settings)
        loaded = load_sheet_rows("overtime", settings)

        self.assertEqual(result["row_count"], 2)
        self.assertEqual(loaded["display_name"], "加班表")
        self.assertEqual(loaded["rows"], [["日期", "開工"], ["2026-05-23", "09:00"]])

    def test_roster_map_prefers_maintenance_copy_over_workbook(self):
        settings = get_settings()
        wb = Workbook()
        wb.active.title = settings.sheets.roster
        ws = wb[settings.sheets.roster]
        ws.cell(1, 1).value = "2026年5月 1 SB"

        save_sheet_rows(
            "roster",
            [
                ["2026年5月 1 SB"],
                ["2026年6月 1 VPP 2 WL21"],
            ],
            settings,
        )

        roster = load_roster_map(settings, wb)

        self.assertEqual(roster[(2026, 6)].day_to_code, {1: "VPP", 2: "WL21"})

    def test_roster_code_definitions_bootstrap_from_roster_cd_columns(self):
        settings = get_settings()
        wb = Workbook()
        wb.active.title = settings.sheets.roster
        ws = wb[settings.sheets.roster]
        ws.cell(1, 3).value = "更碼"
        ws.cell(1, 4).value = "定義"
        ws.cell(2, 3).value = "WL*"
        ws.cell(2, 4).value = "週假"
        ws.cell(3, 3).value = "SB"
        ws.cell(3, 4).value = "Stand by"

        bootstrap_roster_code_definitions(settings, wb)

        self.assertEqual(
            load_roster_code_definitions(settings),
            [
                {"pattern": "WL*", "label": "週假", "sort_order": 1},
                {"pattern": "SB", "label": "Stand by", "sort_order": 2},
            ],
        )

    def test_save_roster_code_definitions_replaces_rows(self):
        settings = get_settings()

        saved = save_roster_code_definitions(
            [{"pattern": "AL*", "label": "Annual leave"}],
            settings,
        )

        self.assertEqual(saved, [{"pattern": "AL*", "label": "Annual leave", "sort_order": 1}])

    def test_import_runtime_inputs_only_requires_roster_and_overtime_sheets(self):
        settings = get_settings()
        wb = Workbook()
        wb.active.title = settings.sheets.roster
        roster_ws = wb[settings.sheets.roster]
        roster_ws.cell(1, 1).value = "2026年5月 1 SB"
        roster_ws.cell(1, 3).value = "更碼"
        roster_ws.cell(1, 4).value = "定義"
        roster_ws.cell(2, 3).value = "SB"
        roster_ws.cell(2, 4).value = "Stand by"
        overtime_ws = wb.create_sheet(settings.sheets.overtime)
        overtime_ws.append(["日期", "開工", "收工"])
        overtime_ws.append(["2026-05-23", "09:00", "18:00"])

        payload = import_runtime_inputs_from_workbook(settings, wb)
        status = list_runtime_input_status(settings)

        self.assertEqual(payload["runtime_input_keys"], ["roster", "overtime"])
        self.assertEqual(load_sheet_rows("roster", settings)["rows"], [["2026年5月 1 SB", None, "更碼", "定義"], [None, None, "SB", "Stand by"]])
        self.assertEqual(load_sheet_rows("overtime", settings)["rows"][1], ["2026-05-23", "09:00", "18:00"])
        self.assertEqual(payload["roster_code_definitions"], [{"pattern": "SB", "label": "Stand by", "sort_order": 1}])
        self.assertEqual(status["excel_role"], "import_only")
        self.assertEqual([sheet["sheet_key"] for sheet in status["sheets"]], ["roster", "overtime"])


if __name__ == "__main__":
    unittest.main()
