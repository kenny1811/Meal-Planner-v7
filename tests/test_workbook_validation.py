import unittest

from openpyxl import Workbook

from meal_planner.excel_io import WorkbookValidationError, header_col_map, validate_workbook_structure
from meal_planner.nutrition_catalog import NUTRIENT_HEADER_BY_KEY
from meal_planner.settings import get_settings


def _make_valid_workbook() -> Workbook:
    settings = get_settings()
    required_sheets = [
        settings.sheets.menu_v5,
        settings.sheets.roster,
        settings.sheets.meal_times,
        settings.sheets.nutrition_list,
        settings.sheets.restaurant,
        settings.sheets.overtime,
        settings.sheets.schedule_grid,
        settings.sheets.payroll_times,
        settings.sheets.public_holidays,
    ]

    wb = Workbook()
    wb.active.title = required_sheets[0]
    for name in required_sheets[1:]:
        wb.create_sheet(name)

    nutrients = list(NUTRIENT_HEADER_BY_KEY.values())
    _write_headers(
        wb[settings.sheets.meal_times],
        ["更碼", "早餐", "午餐", "小食", "晚餐", "餐名", "Pattern"],
    )
    _write_headers(
        wb[settings.sheets.nutrition_list],
        ["類別", "名稱", "暫停", "Min (g)", "Max (g)", "DayMax (g)", *nutrients],
    )
    _write_headers(
        wb[settings.sheets.restaurant],
        ["更碼關鍵字", "舖頭 (Store)", "營業時間", "餐廳選擇", "地址", *nutrients],
    )
    _write_headers(wb[settings.sheets.overtime], ["日期", "開工", "收工"])
    _write_headers(wb[settings.sheets.schedule_grid], ["更碼", "時間", "內容", "時長"])
    return wb


def _write_headers(ws, headers):
    for col, value in enumerate(headers, start=1):
        ws.cell(1, col).value = value


class WorkbookValidationTests(unittest.TestCase):
    def test_valid_workbook_structure_passes(self):
        wb = _make_valid_workbook()

        validate_workbook_structure(wb, get_settings())

    def test_payroll_and_public_holiday_sheets_are_not_core_validation_requirements(self):
        settings = get_settings()
        wb = _make_valid_workbook()
        del wb[settings.sheets.payroll_times]
        del wb[settings.sheets.public_holidays]

        validate_workbook_structure(wb, settings)

    def test_missing_sheet_message_names_sheet_and_existing_sheets(self):
        settings = get_settings()
        wb = _make_valid_workbook()
        del wb[settings.sheets.restaurant]

        with self.assertRaises(WorkbookValidationError) as cm:
            validate_workbook_structure(wb, settings)

        msg = str(cm.exception)
        self.assertIn("工作簿缺少工作表", msg)
        self.assertIn(f"「{settings.sheets.restaurant}」", msg)
        self.assertIn("現有工作表", msg)

    def test_missing_and_duplicate_headers_are_reported_clearly(self):
        settings = get_settings()
        wb = _make_valid_workbook()
        ws = wb[settings.sheets.restaurant]
        ws.cell(1, 2).value = "更碼關鍵字"
        ws.cell(1, 3).value = None

        with self.assertRaises(WorkbookValidationError) as cm:
            validate_workbook_structure(wb, settings)

        msg = str(cm.exception)
        self.assertIn(f"工作表「{settings.sheets.restaurant}」第 1 行有重覆欄名：更碼關鍵字", msg)
        self.assertIn("缺少欄位", msg)
        self.assertIn("舖頭 (Store)", msg)
        self.assertIn("營業時間", msg)

    def test_header_col_map_raises_for_loader_level_missing_column(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "測試"
        _write_headers(ws, ["更碼", "早餐"])

        with self.assertRaises(WorkbookValidationError) as cm:
            header_col_map(ws, 1, required_headers={"更碼", "早餐", "午餐"}, max_scan_col=3)

        self.assertIn("工作表「測試」第 1 行缺少欄位：午餐", str(cm.exception))


if __name__ == "__main__":
    unittest.main()
