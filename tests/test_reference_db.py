import os
import tempfile
import unittest
from datetime import time

from openpyxl import Workbook

from meal_planner.nutrition_catalog import NUTRIENT_HEADER_BY_KEY
from meal_planner.reference_db import load_planning_references
from meal_planner.maintenance_db import save_sheet_rows
from meal_planner.settings import clear_settings_cache, get_settings


def _write_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        ws.cell(row, col).value = value


def _make_workbook() -> Workbook:
    settings = get_settings()
    wb = Workbook()
    wb.active.title = settings.sheets.meal_times
    restaurant_ws = wb.create_sheet(settings.sheets.restaurant)
    schedule_ws = wb.create_sheet(settings.sheets.schedule_grid)

    _write_row(wb[settings.sheets.meal_times], 1, ["更碼", "早餐", "午餐", "小食", "晚餐", None, "餐名", "Pattern"])
    _write_row(wb[settings.sheets.meal_times], 2, ["EleM", "開工前 2 小時", "跟行位表", "跟行位表", "收工後 1.5 小時", None, "早餐", "水果"])
    _write_row(wb[settings.sheets.meal_times], 3, ["其他", "08:00", "12:00", None, None, None, "午餐", "菜+米"])

    nutrients = list(NUTRIENT_HEADER_BY_KEY.values())
    _write_row(restaurant_ws, 1, ["更碼關鍵字", "舖頭 (Store)", "營業時間", "餐廳選擇", "地址", *nutrients])
    _write_row(restaurant_ws, 2, ["Ele*", "店", "09:00-18:00", "餐", "地址", *range(10, 20)])

    _write_row(schedule_ws, 1, ["更碼", "時間", "內容", "時長"])
    _write_row(schedule_ws, 2, ["EleM", time(12, 30), "飯", 45])
    return wb


class ReferenceDatabaseTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_root = os.environ.get("MENU_PROJECT_ROOT")
        os.environ["MENU_PROJECT_ROOT"] = self.tmp.name
        clear_settings_cache()

    def tearDown(self):
        if self.old_root is None:
            os.environ.pop("MENU_PROJECT_ROOT", None)
        else:
            os.environ["MENU_PROJECT_ROOT"] = self.old_root
        clear_settings_cache()
        self.tmp.cleanup()

    def test_references_bootstrap_once_from_workbook(self):
        settings = get_settings()
        wb = _make_workbook()

        first = load_planning_references(settings, wb)
        wb[settings.sheets.meal_times].cell(2, 1).value = "Changed"
        wb[settings.sheets.restaurant].cell(2, 2).value = "改店"
        wb[settings.sheets.schedule_grid].cell(2, 3).value = "Changed"
        second = load_planning_references(settings, wb)

        rules, patterns, restaurants, schedule = first
        self.assertEqual(rules[0].code_pattern, "EleM")
        self.assertEqual(patterns["早餐"], "水果")
        self.assertEqual(restaurants[0]["store"], "店")
        self.assertEqual(restaurants[0]["nutrients"]["kcal"], 10.0)
        self.assertEqual(schedule[0].t, time(12, 30))

        second_rules, _, second_restaurants, second_schedule = second
        self.assertEqual(second_rules[0].code_pattern, "EleM")
        self.assertEqual(second_restaurants[0]["store"], "店")
        self.assertEqual(second_schedule[0].content, "飯")

    def test_schedule_grid_maintenance_rows_override_reference_rows(self):
        settings = get_settings()
        wb = _make_workbook()
        load_planning_references(settings, wb)
        save_sheet_rows(
            "schedule_grid",
            [
                ["更碼", "時間", "內容", "時長", "生效日期"],
                ["EleM", "13:10", "飯", "45", "2026-06-01"],
            ],
            settings,
        )

        _, _, _, schedule = load_planning_references(settings)

        self.assertEqual(schedule[0].t, time(13, 10))
        self.assertEqual(schedule[0].effective_from.isoformat(), "2026-06-01")


if __name__ == "__main__":
    unittest.main()
