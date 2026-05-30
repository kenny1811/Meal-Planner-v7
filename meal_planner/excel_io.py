"""讀取 xlsm（openpyxl）。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from meal_planner.settings import AppSettings, get_settings


class WorkbookValidationError(ValueError):
    """Workbook structure is not usable by the meal planner."""


def _format_sheet_name(name: str) -> str:
    return f"「{name}」"


def _join_items(items: list[str]) -> str:
    return "、".join(items)


def header_col_map(
    ws: Worksheet,
    header_row: int = 1,
    *,
    required_headers: set[str] | None = None,
    max_scan_col: int | None = None,
) -> dict[str, int]:
    out: dict[str, int] = {}
    duplicates: list[str] = []
    scan_to = int(max_scan_col or (ws.max_column or 0))
    for col in range(1, scan_to + 1):
        v = ws.cell(header_row, col).value
        if v is not None and str(v).strip():
            key = str(v).strip()
            if key in out and out[key] != col:
                duplicates.append(f"{key}（第 {out[key]} 欄與第 {col} 欄）")
                continue
            out[key] = col

    errors: list[str] = []
    if duplicates:
        errors.append(
            f"工作表{_format_sheet_name(ws.title)}第 {header_row} 行有重覆欄名："
            f"{_join_items(duplicates)}。"
        )
    if required_headers:
        missing = sorted(required_headers - set(out.keys()))
        if missing:
            errors.append(
                f"工作表{_format_sheet_name(ws.title)}第 {header_row} 行缺少欄位："
                f"{_join_items(missing)}。"
            )
    if errors:
        raise WorkbookValidationError(" ".join(errors))
    return out


def validate_workbook_structure(wb: Workbook, settings: AppSettings | None = None) -> None:
    settings = settings or get_settings()
    sheet_names = {
        "更表": settings.sheets.roster,
        "飯時": settings.sheets.meal_times,
        "餐廳選擇": settings.sheets.restaurant,
        "加班表": settings.sheets.overtime,
        "行位表": settings.sheets.schedule_grid,
    }
    missing_sheets = [name for name in sheet_names.values() if name not in wb.sheetnames]
    errors: list[str] = []
    if missing_sheets:
        errors.append(
            "工作簿缺少工作表："
            f"{_join_items([_format_sheet_name(x) for x in missing_sheets])}。"
        )

    if errors:
        raise WorkbookValidationError(_build_validation_message(errors, wb.sheetnames))

    from meal_planner.nutrition_catalog import NUTRIENT_HEADER_BY_KEY

    nutrient_headers = set(NUTRIENT_HEADER_BY_KEY.values())
    table_specs: list[tuple[str, set[str], int]] = [
        (
            settings.sheets.meal_times,
            {"更碼", "早餐", "午餐", "小食", "晚餐", "餐名", "Pattern"},
            8,
        ),
        (
            settings.sheets.restaurant,
            {"更碼關鍵字", "舖頭 (Store)", "營業時間", "餐廳選擇", "地址"} | nutrient_headers,
            30,
        ),
        (settings.sheets.schedule_grid, {"更碼", "時間", "內容", "時長"}, 12),
        (settings.sheets.overtime, {"日期", "開工", "收工"}, 8),
    ]
    for sheet_name, required, max_col in table_specs:
        try:
            header_col_map(wb[sheet_name], 1, required_headers=required, max_scan_col=max_col)
        except WorkbookValidationError as e:
            errors.append(str(e))

    if errors:
        raise WorkbookValidationError(_build_validation_message(errors, wb.sheetnames))


def _build_validation_message(errors: list[str], sheetnames: list[str]) -> str:
    body = "\n".join(f"- {e}" for e in errors)
    available = _join_items([_format_sheet_name(x) for x in sheetnames])
    return f"Excel 結構有問題，請先修正工作簿：\n{body}\n現有工作表：{available}"


def load_workbook_data(path: Path | None = None, *, validate: bool = True) -> Workbook:
    p = path or get_settings().workbook_path
    # 唔用 read_only：要跨多個 sheet 隨機讀儲存格，read_only 會觸發 ZIP 已關閉錯誤。
    wb = load_workbook(filename=p, read_only=False, data_only=True)
    if validate:
        try:
            validate_workbook_structure(wb, get_settings())
        except Exception:
            wb.close()
            raise
    return wb


def get_sheet(wb: Workbook, name: str) -> Worksheet:
    if name not in wb.sheetnames:
        raise WorkbookValidationError(
            "Excel 結構有問題，請先修正工作簿：\n"
            f"- 工作簿缺少工作表：{_format_sheet_name(name)}。\n"
            f"現有工作表：{_join_items([_format_sheet_name(x) for x in wb.sheetnames])}"
        )
    return wb[name]


def iter_sheet_column_a(ws: Worksheet, max_row: int | None = None) -> Iterator[Any]:
    limit = max_row or (ws.max_row or 0)
    for row in range(1, limit + 1):
        yield ws.cell(row=row, column=1).value


def read_menu_v5_indicators(settings: AppSettings, wb: Workbook) -> tuple[list[str], list[Any], list[Any]]:
    """
    回傳 (營養欄標題列 F.., 返工日列儲存格, 非返工日列儲存格)。
    """
    ws = get_sheet(wb, settings.sheets.menu_v5)
    layout = settings.menu_v5_layout
    h_row = layout.indicator_header_row
    w_row = layout.workday_indicator_row
    nw_row = layout.nonworkday_indicator_row
    c0 = layout.nutrient_first_col
    n = layout.nutrient_col_count
    headers = [ws.cell(row=h_row, column=c0 + i).value for i in range(n)]
    work_vals = [ws.cell(row=w_row, column=c0 + i).value for i in range(n)]
    nonwork_vals = [ws.cell(row=nw_row, column=c0 + i).value for i in range(n)]
    return headers, work_vals, nonwork_vals


def load_roster_map(settings: AppSettings, wb: Workbook | None = None):
    from meal_planner.roster import roster_for_month

    try:
        from meal_planner.maintenance_db import load_sheet_rows

        sheet = load_sheet_rows("roster", settings, wb)
        rows = (
            row[0] if isinstance(row, list) and row else None
            for row in sheet.get("rows", [])
        )
        roster = roster_for_month(rows)
        if roster:
            return roster
    except Exception:
        pass

    if wb is None:
        return {}
    ws = get_sheet(wb, settings.sheets.roster)
    rows = (ws.cell(row=r, column=1).value for r in range(1, (ws.max_row or 0) + 1))
    return roster_for_month(rows)
