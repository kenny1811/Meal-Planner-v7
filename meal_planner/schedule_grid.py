"""行位表：報開工／報收工／飯／小食鐘點；與飯時規則合併成實際用餐時間字串。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from meal_planner.excel_io import get_sheet
from meal_planner.settings import AppSettings

MEAL_KEYS = ("早餐", "午餐", "小食", "晚餐")

_RE_BEFORE = re.compile(r"開工前\s*(\d+(?:\.\d+)?)\s*小時")
_RE_AFTER = re.compile(r"收工後\s*(\d+(?:\.\d+)?)\s*小時")


def _to_time(v: Any) -> time | None:
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                pass
    return None


def _to_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _ddt(d: date, t: time) -> datetime:
    return datetime.combine(d, t)


def _fmt_clock(t: time) -> str:
    return t.strftime("%H:%M")


def grid_row_matches_roster(cell_code: str | None, roster_code: str) -> bool:
    """行位更碼：與更表字串一致；唔匹配「PenC頂位」當「PenC」。"""
    if not cell_code or not roster_code:
        return False
    c = str(cell_code).strip()
    r = str(roster_code).strip()
    if c == r:
        return True
    if c.startswith(r + "(") or c.startswith(r + "（"):
        return True
    return False


def _header_col_map(
    ws: Worksheet,
    header_row: int = 1,
    *,
    required_headers: set[str] | None = None,
    max_scan_col: int | None = None,
) -> dict[str, int]:
    out: dict[str, int] = {}
    scan_to = int(max_scan_col or (ws.max_column or 0))
    for col in range(1, scan_to + 1):
        v = ws.cell(header_row, col).value
        if v is not None and str(v).strip():
            key = str(v).strip()
            if key in out and out[key] != col:
                raise ValueError(
                    f"工作表「{ws.title}」第 {header_row} 行有重覆欄名「{key}」"
                    f"（第 {out[key]} 欄與第 {col} 欄）。"
                )
            out[key] = col
    return out


@dataclass
class ScheduleRow:
    code: str
    t: time | None
    content: str
    duration_min: int | None
    effective_from: date | None = None


def load_schedule_rows(ws: Worksheet) -> list[ScheduleRow]:
    h = _header_col_map(
        ws,
        1,
        required_headers={"更碼", "時間", "內容", "時長"},
        max_scan_col=12,
    )
    c_code = h.get("更碼")
    c_time = h.get("時間")
    c_content = h.get("內容")
    c_dur = h.get("時長")
    c_eff = h.get("生效日期") or h.get("生效") or h.get("Effective From")
    if not c_code or not c_time or not c_content:
        return []
    rows: list[ScheduleRow] = []
    for r in range(2, (ws.max_row or 0) + 1):
        code = ws.cell(r, c_code).value
        if code is None or str(code).strip() == "":
            continue
        dur_raw = ws.cell(r, c_dur).value if c_dur else None
        dur: int | None = None
        if dur_raw is not None:
            try:
                dur = int(float(dur_raw))
            except (TypeError, ValueError):
                dur = None
        rows.append(
            ScheduleRow(
                code=str(code).strip(),
                t=_to_time(ws.cell(r, c_time).value),
                content=str(ws.cell(r, c_content).value or ""),
                duration_min=dur,
                effective_from=_to_date(ws.cell(r, c_eff).value) if c_eff else None,
            )
        )
    return rows


def rows_for_roster(rows: list[ScheduleRow], roster_code: str, day: date | None = None) -> list[ScheduleRow]:
    matched = [x for x in rows if grid_row_matches_roster(x.code, roster_code)]
    if day is None:
        return matched
    dated = [x for x in matched if x.effective_from is not None and x.effective_from <= day]
    if not dated:
        return [x for x in matched if x.effective_from is None]
    latest = max(x.effective_from for x in dated if x.effective_from is not None)
    return [x for x in dated if x.effective_from == latest]


def load_schedule_rows_from_rows(rows: list[list[Any]]) -> list[ScheduleRow]:
    if not rows:
        return []
    headers = {str(v).strip(): idx for idx, v in enumerate(rows[0]) if v is not None and str(v).strip()}
    c_code = headers.get("更碼")
    c_time = headers.get("時間")
    c_content = headers.get("內容")
    c_dur = headers.get("時長")
    c_eff = headers.get("生效日期")
    if c_eff is None:
        c_eff = headers.get("生效")
    if c_eff is None:
        c_eff = headers.get("Effective From")
    if c_code is None or c_time is None or c_content is None:
        return []
    out: list[ScheduleRow] = []
    for row in rows[1:]:
        if not isinstance(row, list) or c_code >= len(row):
            continue
        code = str(row[c_code] or "").strip()
        if not code:
            continue
        dur_raw = row[c_dur] if c_dur is not None and c_dur < len(row) else None
        dur: int | None = None
        if dur_raw is not None:
            try:
                dur = int(float(dur_raw))
            except (TypeError, ValueError):
                dur = None
        out.append(
            ScheduleRow(
                code=code,
                t=_to_time(row[c_time]) if c_time < len(row) else None,
                content=str(row[c_content] if c_content < len(row) and row[c_content] is not None else ""),
                duration_min=dur,
                effective_from=_to_date(row[c_eff]) if c_eff is not None and c_eff < len(row) else None,
            )
        )
    return out


def report_start_end(rows: list[ScheduleRow]) -> tuple[time | None, time | None]:
    start: time | None = None
    end: time | None = None
    for x in rows:
        if "報開工" in x.content and x.t is not None:
            start = x.t
            break
    for x in reversed(rows):
        if "報收工" in x.content and x.t is not None:
            end = x.t
            break
    return start, end


def first_food_time(rows: list[ScheduleRow], *, keyword: str) -> tuple[time | None, int | None]:
    """keyword 為「飯」或「小食」（內容包含即計）。"""
    for x in rows:
        if keyword not in x.content:
            continue
        if x.t is None:
            continue
        return x.t, x.duration_min
    return None, None


def first_snack_time(rows: list[ScheduleRow]) -> tuple[time | None, int | None]:
    """
    小食時間：只接受內容包含「小食」。
    """
    for x in rows:
        txt = x.content or ""
        if x.t is None:
            continue
        if "小食" in txt:
            return x.t, x.duration_min
    return None, None


def overtime_override(ws: Worksheet, day: date) -> tuple[time | None, time | None]:
    h = _header_col_map(
        ws,
        1,
        required_headers={"日期", "開工", "收工"},
        max_scan_col=8,
    )
    c_date = h.get("日期")
    c_start = h.get("開工")
    c_end = h.get("收工")
    if not c_date:
        return None, None
    for r in range(2, (ws.max_row or 0) + 1):
        dv = ws.cell(r, c_date).value
        dd: date | None = None
        if isinstance(dv, datetime):
            dd = dv.date()
        elif isinstance(dv, date):
            dd = dv
        if dd != day:
            continue
        return _to_time(ws.cell(r, c_start).value) if c_start else None, _to_time(ws.cell(r, c_end).value) if c_end else None
    return None, None


def load_overtime_overrides(ws: Worksheet) -> dict[date, tuple[time | None, time | None]]:
    h = _header_col_map(
        ws,
        1,
        required_headers={"日期", "開工", "收工"},
        max_scan_col=8,
    )
    c_date = h.get("日期")
    c_start = h.get("開工")
    c_end = h.get("收工")
    if not c_date:
        return {}
    out: dict[date, tuple[time | None, time | None]] = {}
    for r in range(2, (ws.max_row or 0) + 1):
        dv = ws.cell(r, c_date).value
        dd: date | None = None
        if isinstance(dv, datetime):
            dd = dv.date()
        elif isinstance(dv, date):
            dd = dv
        if dd is None:
            continue
        out[dd] = (
            _to_time(ws.cell(r, c_start).value) if c_start else None,
            _to_time(ws.cell(r, c_end).value) if c_end else None,
        )
    return out


def load_overtime_overrides_from_rows(rows: list[list[Any]]) -> dict[date, tuple[time | None, time | None]]:
    if not rows:
        return {}
    headers = {str(v).strip(): idx for idx, v in enumerate(rows[0]) if v is not None and str(v).strip()}
    c_date = headers.get("日期")
    c_start = headers.get("開工")
    c_end = headers.get("收工")
    if c_date is None:
        return {}
    out: dict[date, tuple[time | None, time | None]] = {}
    for row in rows[1:]:
        if not isinstance(row, list) or c_date >= len(row):
            continue
        dd = _to_date(row[c_date])
        if dd is None:
            continue
        out[dd] = (
            _to_time(row[c_start]) if c_start is not None and c_start < len(row) else None,
            _to_time(row[c_end]) if c_end is not None and c_end < len(row) else None,
        )
    return out


def parse_relative_hours(text: str | None, *, kind: str) -> float | None:
    if not text:
        return None
    s = str(text).strip()
    if kind == "before":
        m = _RE_BEFORE.search(s)
        return float(m.group(1)) if m else None
    if kind == "after":
        m = _RE_AFTER.search(s)
        return float(m.group(1)) if m else None
    return None


def _cell_rule_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    s = str(v).strip()
    return s if s else None


def resolve_meal_times_display(
    settings: AppSettings,
    wb: Workbook | None,
    *,
    day: date,
    roster_code: str,
    primary_rule: dict[str, Any] | None,
    is_work_day: bool | None,
    restaurant: dict[str, Any] | None,
    schedule_rows: list[ScheduleRow] | None = None,
    overtime_overrides: dict[date, tuple[time | None, time | None]] | None = None,
) -> dict[str, str | None]:
    """
    回傳 早餐／午餐／小食／晚餐 → 顯示字串（實際鐘点或時段）。
    無法推算則 None（前端可回落顯示飯時原文）。
    """
    out: dict[str, str | None] = {k: None for k in MEAL_KEYS}
    if not primary_rule or not roster_code:
        return out

    if schedule_rows is None:
        if wb is None:
            return out
        try:
            sg_ws = get_sheet(wb, settings.sheets.schedule_grid)
        except KeyError:
            return out
        all_rows = load_schedule_rows(sg_ws)
    else:
        all_rows = schedule_rows
    my_rows = rows_for_roster(all_rows, roster_code, day)
    g_start, g_end = report_start_end(my_rows)

    ot_start, ot_end = None, None
    if overtime_overrides is not None:
        ot_start, ot_end = overtime_overrides.get(day, (None, None))
    else:
        try:
            if wb is None:
                raise KeyError
            ot_ws = get_sheet(wb, settings.sheets.overtime)
            ot_start, ot_end = overtime_override(ot_ws, day)
        except KeyError:
            pass

    start = ot_start if ot_start is not None else g_start
    end = ot_end if ot_end is not None else g_end

    def resolve_one(meal: str, raw: Any) -> str | None:
        s = _cell_rule_str(raw)
        if not s or s == "—":
            return None
        if re.fullmatch(r"\d{1,2}:\d{2}", s):
            return s

        if "開工前" in s and start is not None:
            h = parse_relative_hours(s, kind="before") or 2.0
            a = _ddt(day, start) - timedelta(hours=h)
            return _fmt_clock(a.time())

        if "收工後" in s and end is not None:
            h = parse_relative_hours(s, kind="after") or 1.5
            b = _ddt(day, end) + timedelta(hours=h)
            return _fmt_clock(b.time())

        if "跟行位表" in s:
            if meal == "午餐":
                t, _ = first_food_time(my_rows, keyword="飯")
                if t is None:
                    return None
                return _fmt_clock(t)
            if meal == "小食":
                t, _ = first_snack_time(my_rows)
                if t is None:
                    return None
                return _fmt_clock(t)
            return None

        return None

    for meal in MEAL_KEYS:
        raw = primary_rule.get(meal)
        out[meal] = resolve_one(meal, raw)

    return out
