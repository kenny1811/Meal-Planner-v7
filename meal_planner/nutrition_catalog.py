"""營養清單查找：按 §11.2 規則由 Pattern item 對應食材列。"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from meal_planner.excel_io import header_col_map
from meal_planner.indicators import NUTRIENT_KEYS

NUTRIENT_HEADER_BY_KEY = {
    "kcal": "卡路里 (kCal)",
    "protein_g": "蛋白質 (g)",
    "carb_g": "碳水 (g)",
    "sugar_g": "天然糖 (g)",
    "cholesterol_mg": "膽固醇 (mg)",
    "sodium_mg": "鈉 (mg)",
    "calcium_mg": "鈣 (mg)",
    "fat_total_g": "總脂肪 (g)",
    "fat_sat_g": "飽和脂肪 (g)",
    "fat_trans_g": "反式脂肪 (g)",
}


@dataclass(frozen=True)
class NutritionEntry:
    row_index: int
    paused: bool
    category: str
    name: str
    nutrients: dict[str, float]
    min_g: float | None
    max_g: float | None
    daymax_g: float | None


def _is_paused(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"yes", "y", "true", "1"}


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_float_or_none(v: Any) -> float | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_nutrition_entries(ws: Worksheet) -> list[NutritionEntry]:
    required = {"類別", "名稱", "暫停", "Min (g)", "Max (g)", "DayMax (g)"}
    required.update(NUTRIENT_HEADER_BY_KEY.values())
    h = header_col_map(ws, 1, required_headers=required, max_scan_col=20)
    c_pause = h.get("暫停")
    c_cat = h.get("類別")
    c_name = h.get("名稱")
    c_min = h.get("Min (g)")
    c_max = h.get("Max (g)")
    c_daymax = h.get("DayMax (g)")
    if not c_cat or not c_name:
        return []
    out: list[NutritionEntry] = []
    for r in range(2, (ws.max_row or 0) + 1):
        cat = ws.cell(r, c_cat).value
        name = ws.cell(r, c_name).value
        if cat is None or name is None:
            continue
        cat_s = str(cat).strip()
        name_s = str(name).strip()
        if not cat_s or not name_s:
            continue
        paused = _is_paused(ws.cell(r, c_pause).value) if c_pause else False
        nutrients: dict[str, float] = {}
        for key in NUTRIENT_KEYS:
            col = h.get(NUTRIENT_HEADER_BY_KEY[key])
            nutrients[key] = _to_float(ws.cell(r, col).value) if col else 0.0
        out.append(
            NutritionEntry(
                row_index=r,
                paused=paused,
                category=cat_s,
                name=name_s,
                nutrients=nutrients,
                min_g=_to_float_or_none(ws.cell(r, c_min).value) if c_min else None,
                max_g=_to_float_or_none(ws.cell(r, c_max).value) if c_max else None,
                daymax_g=_to_float_or_none(ws.cell(r, c_daymax).value) if c_daymax else None,
            )
        )
    return out


def _match_entries_for_token(entries: list[NutritionEntry], token: str) -> list[NutritionEntry]:
    t = token.strip().lower()
    if not t:
        return []

    exact_cat = [e for e in entries if e.category.lower() == t and not e.paused]
    if exact_cat:
        return exact_cat

    by_name = [e for e in entries if t in e.name.lower() and not e.paused]
    if by_name:
        return by_name
    return []


def resolve_item_name_from_alternatives(
    entries: list[NutritionEntry],
    alternatives: list[str],
) -> str | None:
    """
    item 內 alternatives 由左到右嘗試；每個 token 先類別 exact，再名稱 contains。
    一旦找到候選，回傳候選中第一個名稱（保留工作簿原順序）。
    """
    for token in alternatives:
        matches = _match_entries_for_token(entries, token)
        if matches:
            return matches[0].name
    return None


def resolve_item_entry_from_alternatives(
    entries: list[NutritionEntry],
    alternatives: list[str],
) -> NutritionEntry | None:
    for token in alternatives:
        matches = _match_entries_for_token(entries, token)
        if matches:
            return matches[0]
    return None


def candidate_entries_from_alternatives(
    entries: list[NutritionEntry],
    alternatives: list[str],
) -> list[NutritionEntry]:
    """
    item 內 `/` 候選：按 alternatives 次序串接候選清單（去重，保留原順序）。
    """
    out: list[NutritionEntry] = []
    seen_rows: set[int] = set()
    for token in alternatives:
        matches = _match_entries_for_token(entries, token)
        for e in matches:
            if e.row_index in seen_rows:
                continue
            out.append(e)
            seen_rows.add(e.row_index)
    return out

