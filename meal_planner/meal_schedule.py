"""飯時、餐廳選擇：更碼匹配（含 wildcard）與每日餐單草稿欄位。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Any

from openpyxl.workbook.workbook import Workbook

from meal_planner.excel_io import get_sheet, header_col_map
from meal_planner.indicators import DayIndicatorProfile, NUTRIENT_KEYS
from meal_planner.nutrition_catalog import (
    NUTRIENT_HEADER_BY_KEY,
    candidate_entries_from_alternatives,
)
from meal_planner.nutrition_db import load_catalog_entries
from meal_planner.optimizer import solve_day_meal_plan
from meal_planner.patterns import parse_meal_patterns
from meal_planner.settings import AppSettings, MealTimesStackConfig


def roster_matches_rule(rule_cell: str | None, roster_code: str) -> bool:
    if not rule_cell or not roster_code:
        return False
    rule = str(rule_cell).strip()
    code = roster_code.strip()
    if not rule:
        return False
    if rule == "其他":
        return False
    rule_cmp = rule.casefold()
    code_cmp = code.casefold()
    if rule_cmp.endswith("*"):
        return code_cmp.startswith(rule_cmp[:-1])
    return code_cmp == rule_cmp


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    s = str(v).strip()
    return s if s else None


@dataclass(frozen=True)
class MealTimeRule:
    row_index: int
    code_pattern: str
    breakfast: str | None
    lunch: str | None
    snack: str | None
    dinner: str | None


@dataclass(frozen=True)
class MealPlanningCache:
    meal_time_rules: list[MealTimeRule]
    meal_patterns: dict[str, str | None]
    restaurant_rows: list[dict[str, Any]]
    nutrition_entries: list[Any]
    schedule_rows: list[Any]
    overtime_overrides: dict[date, tuple[time | None, time | None]]


def load_meal_time_rules(ws: Worksheet) -> list[MealTimeRule]:
    h = header_col_map(
        ws,
        1,
        required_headers={"更碼", "早餐", "午餐", "小食", "晚餐"},
        max_scan_col=8,  # 只讀左側時間主表 A:E（含少量緩衝）
    )
    c_code = h.get("更碼")
    if not c_code:
        return []
    keys = ("早餐", "午餐", "小食", "晚餐")
    cols = {k: h.get(k) for k in keys}
    rules: list[MealTimeRule] = []
    for r in range(2, (ws.max_row or 0) + 1):
        code = ws.cell(r, c_code).value
        if code is None or str(code).strip() == "":
            continue
        rules.append(
            MealTimeRule(
                row_index=r,
                code_pattern=str(code).strip(),
                breakfast=_cell_str(ws.cell(r, cols["早餐"]).value) if cols["早餐"] else None,
                lunch=_cell_str(ws.cell(r, cols["午餐"]).value) if cols["午餐"] else None,
                snack=_cell_str(ws.cell(r, cols["小食"]).value) if cols["小食"] else None,
                dinner=_cell_str(ws.cell(r, cols["晚餐"]).value) if cols["晚餐"] else None,
            )
        )
    return rules


def load_meal_patterns_table(ws: Worksheet) -> dict[str, str | None]:
    """
    讀取飯時表 G:H（餐名/Pattern）作為獨立 pattern 表。
    與 A:E（更碼/時間）完全分開，不按同一行綁定。
    """
    h = header_col_map(
        ws,
        1,
        required_headers={"餐名", "Pattern"},
        max_scan_col=8,  # 只讀 G:H pattern 表
    )
    c_meal = h.get("餐名")
    c_pat = h.get("Pattern")
    out: dict[str, str | None] = {m: None for m in MEAL_LABELS}
    if not c_meal or not c_pat:
        return out
    for r in range(2, (ws.max_row or 0) + 1):
        meal = _cell_str(ws.cell(r, c_meal).value)
        pat = _cell_str(ws.cell(r, c_pat).value)
        if meal not in out or not pat:
            continue
        if out[meal] is None:
            out[meal] = pat
    return out


def first_matching_meal_rule(rules: list[MealTimeRule], roster_code: str) -> MealTimeRule | None:
    """§10.1：由上而下，更碼第一個命中（不含「其他」列，該列單獨兜底）。"""
    for rule in rules:
        if roster_matches_rule(rule.code_pattern, roster_code):
            return rule
    for rule in rules:
        if rule.code_pattern == "其他":
            return rule
    return None


MEAL_LABELS = ("早餐", "午餐", "小食", "晚餐")
MEAL_ROTATION_OFFSET = {"早餐": 0, "午餐": 1, "小食": 2, "晚餐": 3}


def patterns_peninsula_stack(
    meal_patterns: dict[str, str | None],
    pattern_rules: tuple[str, str, str, str],
) -> dict[str, str | None]:
    """
    與時間表分離後，Pattern 直接來自獨立餐名表（G:H）。
    保留此函數介面，回傳該表的四餐映射。
    """
    return {m: meal_patterns.get(m) for m in MEAL_LABELS}


def _use_peninsula_stack(roster_code: str, cfg: MealTimesStackConfig) -> bool:
    return bool(cfg.enabled and roster_code.startswith(cfg.roster_prefix))


def patterns_by_meal_name(rules: list[MealTimeRule], roster_code: str) -> dict[str, str | None]:
    # Pattern 已由獨立表讀取；此函數保留兼容，預設不由時間表推 Pattern。
    return {k: None for k in MEAL_LABELS}


def default_patterns_by_meal_name(rules: list[MealTimeRule]) -> dict[str, str | None]:
    # Pattern 已由獨立表讀取；此函數保留兼容。
    return {k: None for k in MEAL_LABELS}


def load_restaurant_rows(ws: Worksheet) -> list[dict[str, Any]]:
    required = {"更碼關鍵字", "舖頭 (Store)", "營業時間", "餐廳選擇", "地址"}
    required.update(NUTRIENT_HEADER_BY_KEY.values())
    h = header_col_map(ws, 1, required_headers=required, max_scan_col=30)
    c_kw = h.get("更碼關鍵字")
    if not c_kw:
        return []
    cols = {name: h.get(name) for name in ("舖頭 (Store)", "營業時間", "餐廳選擇", "地址")}
    nutrient_cols = {k: h.get(NUTRIENT_HEADER_BY_KEY[k]) for k in NUTRIENT_KEYS}
    rows: list[dict[str, Any]] = []
    for r in range(2, (ws.max_row or 0) + 1):
        kw = ws.cell(r, c_kw).value
        if kw is None or str(kw).strip() == "":
            continue
        def gv(name: str) -> Any:
            ci = cols.get(name)
            return ws.cell(r, ci).value if ci else None

        rows.append(
            {
                "row": r,
                "keyword": str(kw).strip(),
                "store": gv("舖頭 (Store)"),
                "hours": gv("營業時間"),
                "choice": gv("餐廳選擇"),
                "address": gv("地址"),
                "nutrients": {
                    k: float(ws.cell(r, nutrient_cols[k]).value or 0.0) if nutrient_cols[k] else 0.0
                    for k in NUTRIENT_KEYS
                },
            }
        )
    return rows


def build_meal_planning_cache(settings: AppSettings, wb: Workbook | None = None) -> MealPlanningCache:
    from meal_planner.reference_db import load_planning_references

    rules, patterns, restaurant_rows, schedule_rows = load_planning_references(settings, wb)

    nutrition_entries = load_catalog_entries(settings, wb)

    overtime_overrides: dict[date, tuple[time | None, time | None]] = {}
    try:
        from meal_planner.maintenance_db import load_sheet_rows
        from meal_planner.schedule_grid import load_overtime_overrides_from_rows

        overtime_sheet = load_sheet_rows("overtime", settings, wb)
        overtime_overrides = load_overtime_overrides_from_rows(overtime_sheet.get("rows", []))
    except Exception:
        if wb is not None:
            try:
                from meal_planner.schedule_grid import load_overtime_overrides

                overtime_overrides = load_overtime_overrides(get_sheet(wb, settings.sheets.overtime))
            except KeyError:
                pass

    return MealPlanningCache(
        meal_time_rules=rules,
        meal_patterns=patterns,
        restaurant_rows=restaurant_rows,
        nutrition_entries=nutrition_entries,
        schedule_rows=schedule_rows,
        overtime_overrides=overtime_overrides,
    )


def first_matching_restaurant(rest_rows: list[dict[str, Any]], roster_code: str) -> dict[str, Any] | None:
    for row in rest_rows:
        if roster_matches_rule(row["keyword"], roster_code):
            return row
    return None


def choose_ingredients_for_meals(
    wb: Workbook | None,
    settings: AppSettings,
    meal_pattern_parts: dict[str, list[dict[str, object]]],
    day: date | None = None,
    indicators: DayIndicatorProfile | None = None,
    visible_meals: set[str] | None = None,
    fixed_nutrients: dict[str, float] | None = None,
    fixed_meals: set[str] | None = None,
    reroll_nonce: int = 0,
    nutrition_entries: list[Any] | None = None,
) -> tuple[
    dict[str, list[str]],
    dict[str, dict[str, float]],
    dict[str, list[dict[str, object]]],
    dict[str, Any],
]:
    """
    依 §11.2：每個 item 先類別 exact，再名稱 contains；左右候選取第一個可用。
    回傳每餐對應食材名稱列表（按 pattern item 次序）。
    """
    def default_grams_for_entry(entry: Any) -> float:
        if entry.min_g is not None:
            return float(round(float(entry.min_g)))
        return 0.0

    if nutrition_entries is None:
        entries = load_catalog_entries(settings, wb)
    else:
        entries = nutrition_entries
    visible_set = set(visible_meals or meal_pattern_parts.keys())
    fixed_meal_set = set(fixed_meals or set())
    rice_token = settings.rice.rice_category_exact.strip().lower()
    day_offset = (int(day.day) if isinstance(day, date) else 0) + int(reroll_nonce or 0)

    candidates_by_item: dict[tuple[str, int], list[Any]] = {}
    for meal, items in meal_pattern_parts.items():
        if meal not in visible_set:
            continue
        for i, item in enumerate(items):
            alts = item.get("alternatives", [])
            alts_list = [str(x) for x in alts] if isinstance(alts, list) else []
            candidates_by_item[(meal, i)] = candidate_entries_from_alternatives(entries, alts_list)

    fixed_names: dict[str, list[str]] = {meal: [] for meal in meal_pattern_parts.keys()}
    fixed_items: dict[str, list[dict[str, object]]] = {meal: [] for meal in meal_pattern_parts.keys()}
    fixed_meal_nutrients: dict[str, dict[str, float]] = {
        meal: {k: 0.0 for k in NUTRIENT_KEYS}
        for meal in meal_pattern_parts.keys()
    }
    extra_fixed_nutrients = {k: float((fixed_nutrients or {}).get(k, 0.0) or 0.0) for k in NUTRIENT_KEYS}
    for meal in fixed_meal_set:
        if meal not in visible_set:
            continue
        for i, item in enumerate(meal_pattern_parts.get(meal, [])):
            candidates = candidates_by_item.get((meal, i), [])
            entry = None
            if candidates:
                meal_offset = MEAL_ROTATION_OFFSET.get(meal, 0)
                entry = candidates[(day_offset + meal_offset + i) % len(candidates)]
            if entry is None:
                raw = str(item.get("raw", "")).strip()
                if raw:
                    fixed_names[meal].append(raw)
                    fixed_items[meal].append({"name": raw, "grams": None, "row": None})
                continue
            grams = default_grams_for_entry(entry)
            fixed_names[meal].append(f"{entry.name}({grams:.0f}g)")
            fixed_items[meal].append({"name": entry.name, "grams": grams, "row": entry.row_index})
            ratio = grams / 100.0
            for k in NUTRIENT_KEYS:
                v = float(entry.nutrients.get(k, 0.0)) * ratio
                fixed_meal_nutrients[meal][k] += v
                extra_fixed_nutrients[k] += v

    solver_visible_set = visible_set - fixed_meal_set

    if indicators is not None:
        solved = solve_day_meal_plan(
            settings=settings,
            indicators=indicators,
            meal_pattern_parts=meal_pattern_parts,
            candidates_by_item=candidates_by_item,
            visible_meals=solver_visible_set,
            rice_token=rice_token,
            day_offset=day_offset,
            reroll_nonce=int(reroll_nonce or 0),
            fixed_nutrients=extra_fixed_nutrients,
        )
        if solved is not None:
            for meal in fixed_meal_set:
                if meal in visible_set:
                    solved.meal_ingredients[meal] = fixed_names.get(meal, [])
                    solved.meal_nutrients[meal] = fixed_meal_nutrients.get(meal, {k: 0.0 for k in NUTRIENT_KEYS})
                    solved.meal_items[meal] = fixed_items.get(meal, [])
            return (
                solved.meal_ingredients,
                solved.meal_nutrients,
                solved.meal_items,
                {"mode": "milp", "status": solved.status, **solved.diagnostics},
            )

    out_names: dict[str, list[str]] = {}
    out_nutrients: dict[str, dict[str, float]] = {}
    out_items: dict[str, list[dict[str, object]]] = {}
    pick_cursor: dict[str, int] = {}
    rice_locked_entry = None
    for meal, items in meal_pattern_parts.items():
        if meal in fixed_meal_set and meal in visible_set:
            out_names[meal] = fixed_names.get(meal, [])
            out_nutrients[meal] = fixed_meal_nutrients.get(meal, {k: 0.0 for k in NUTRIENT_KEYS})
            out_items[meal] = fixed_items.get(meal, [])
            continue
        if meal not in visible_set:
            out_names[meal] = []
            out_nutrients[meal] = {k: 0.0 for k in NUTRIENT_KEYS}
            out_items[meal] = []
            continue
        chosen: list[str] = []
        nutrient_sum = {k: 0.0 for k in NUTRIENT_KEYS}
        chosen_items: list[dict[str, object]] = []
        for i, item in enumerate(items):
            alts = item.get("alternatives", [])
            alts_list = [str(x) for x in alts] if isinstance(alts, list) else []
            candidates = candidates_by_item.get((meal, i), [])
            entry = None
            is_rice_item = any((a or "").strip().lower() == rice_token for a in alts_list)
            if is_rice_item:
                if rice_locked_entry is None and candidates:
                    key = "|".join(alts_list)
                    base = pick_cursor.get(key, 0)
                    idx = (base + day_offset) % len(candidates)
                    rice_locked_entry = candidates[idx]
                    pick_cursor[key] = base + 1
                entry = rice_locked_entry
            elif candidates:
                # 輪替按「餐次 + alternatives」分開，避免午餐序列被晚餐推進。
                key = f"{meal}|{'|'.join(alts_list)}"
                base = pick_cursor.get(key, 0)
                meal_offset = MEAL_ROTATION_OFFSET.get(meal, 0)
                idx = (base + day_offset + meal_offset) % len(candidates)
                entry = candidates[idx]
                pick_cursor[key] = base + 1
            if entry is not None:
                # Min(g) 有值就照用（包括 0）；Min 空白時用 default_g，但不可超過 Max/DayMax。
                grams = default_grams_for_entry(entry)
                chosen.append(f"{entry.name}({grams:.0f}g)")
                chosen_items.append({"name": entry.name, "grams": grams, "row": entry.row_index})
                ratio = grams / 100.0
                for k in NUTRIENT_KEYS:
                    nutrient_sum[k] += float(entry.nutrients.get(k, 0.0)) * ratio
            else:
                raw = str(item.get("raw", "")).strip()
                if raw:
                    chosen.append(raw)
                    chosen_items.append({"name": raw, "grams": None, "row": None})
        out_names[meal] = chosen
        out_nutrients[meal] = nutrient_sum
        out_items[meal] = chosen_items
    return out_names, out_nutrients, out_items, {"mode": "fallback_rotation"}


def build_rice_note(
    meal_items: dict[str, list[dict[str, object]]],
    settings: AppSettings,
    visible_meals: set[str] | None = None,
) -> str:
    """
    依 §13：同日只一款米。根據已選米類熟重總和，換算生重與水重。
    """
    rice_items: list[tuple[str, float]] = []
    for meal, items in meal_items.items():
        if visible_meals is not None and meal not in visible_meals:
            continue
        if not isinstance(items, list):
            continue
        for it in items:
            name = str(it.get("name", "")).strip()
            grams = it.get("grams")
            if not name or not isinstance(grams, (int, float)):
                continue
            rice_markers = tuple(x for x in settings.rice.note_name_contains if x)
            if any(marker in name for marker in rice_markers):
                rice_items.append((name, float(grams)))

    if not rice_items:
        return "（米類備註：配餐後填）"

    rice_name = rice_items[0][0]
    cooked_g = sum(g for _, g in rice_items)
    is_brown = bool(settings.rice.brown_name_contains and settings.rice.brown_name_contains in rice_name)
    ratio = settings.rice.cooked_to_raw_brown if is_brown else settings.rice.cooked_to_raw_other
    raw_g = cooked_g / ratio if ratio > 0 else 0.0
    water_g = raw_g * settings.rice.water_multiplier
    return f"{rice_name}({cooked_g:.0f}g)=生重{raw_g:.0f}g\n水={water_g:.0f}g"


def build_day_meal_plan(
    settings: AppSettings,
    wb: Workbook | None,
    roster_code: str | None,
    is_work_day: bool | None,
    day: date | None = None,
    indicators: DayIndicatorProfile | None = None,
    reroll_nonce: int = 0,
    cache: MealPlanningCache | None = None,
) -> dict[str, Any]:
    """組合飯時主規則、各餐 Pattern、返工日午餐餐廳（第一命中）；可選 `day` 以解析行位表實際用餐時間。"""
    if not roster_code:
        return {
            "primary_rule": None,
            "meal_patterns": {},
            "restaurant_lunch": None,
            "meal_times_resolved": {},
            "note": "無更表更碼，無法對應飯時。",
        }

    rules = cache.meal_time_rules if cache is not None else load_meal_time_rules(get_sheet(wb, settings.sheets.meal_times))
    primary = first_matching_meal_rule(rules, roster_code)
    pattern_table = cache.meal_patterns if cache is not None else load_meal_patterns_table(get_sheet(wb, settings.sheets.meal_times))

    rest = None
    if (not settings.meal_business_rules.restaurant_lunch_workday_only) or is_work_day is True:
        r_rows = cache.restaurant_rows if cache is not None else load_restaurant_rows(get_sheet(wb, settings.sheets.restaurant))
        hit = first_matching_restaurant(r_rows, roster_code)
        if hit:
            rest = {
                "keyword": hit["keyword"],
                "store": hit["store"],
                "hours": hit["hours"],
                "choice": hit["choice"],
                "address": hit["address"],
                "nutrients": hit.get("nutrients", {}),
            }

    primary_dict = None
    if primary:
        primary_dict = {
            "code_pattern": primary.code_pattern,
            "早餐": primary.breakfast,
            "午餐": primary.lunch,
            "小食": primary.snack,
            "晚餐": primary.dinner,
            "餐名": None,
            "pattern": None,
        }

    meal_times_resolved: dict[str, Any] = {}
    if day is not None and primary_dict:
        from meal_planner.schedule_grid import resolve_meal_times_display

        meal_times_resolved = resolve_meal_times_display(
            settings,
            wb,
            day=day,
            roster_code=roster_code,
            primary_rule=primary_dict,
            is_work_day=is_work_day,
            restaurant=rest,
            schedule_rows=cache.schedule_rows if cache is not None else None,
            overtime_overrides=cache.overtime_overrides if cache is not None else None,
        )
    if meal_times_resolved:
        visible_meals = {
            meal
            for meal in MEAL_LABELS
            if (
                meal_times_resolved.get(meal) is not None
                and str(meal_times_resolved.get(meal)).strip() != ""
            )
        }
    else:
        # 無 day/無 resolved 時，退回 A:E 命中行本身嘅時間欄判斷可見餐次
        visible_meals = set()
        if primary is not None:
            if primary.breakfast:
                visible_meals.add("早餐")
            if primary.lunch:
                visible_meals.add("午餐")
            if primary.snack:
                visible_meals.add("小食")
            if primary.dinner:
                visible_meals.add("晚餐")

    # 規則：先由 A:E 決定有邊幾餐；再只為該幾餐從 G:H 拿 Pattern。
    # 餐廳午餐是固定營養值，求解時要當成已食用，其他餐次再遷就它。
    fixed_nutrients = None
    solver_visible_meals = set(visible_meals)
    if "午餐" in visible_meals and rest and isinstance(rest.get("nutrients"), dict):
        fixed_nutrients = {
            k: float(rest["nutrients"].get(k, 0.0) or 0.0)
            for k in NUTRIENT_KEYS
        }
        solver_visible_meals.discard("午餐")
    by_meal = {m: (pattern_table.get(m) if m in solver_visible_meals else None) for m in MEAL_LABELS}
    fixed_meals = {
        meal
        for meal in settings.meal_business_rules.fixed_meals
        if meal in solver_visible_meals and pattern_table.get(meal)
    }

    meal_pattern_parts = parse_meal_patterns(by_meal, settings.pattern)
    meal_ingredients, meal_nutrients, meal_items, optimization_meta = choose_ingredients_for_meals(
        wb,
        settings,
        meal_pattern_parts,
        day=day,
        indicators=indicators,
        visible_meals=solver_visible_meals,
        fixed_nutrients=fixed_nutrients,
        fixed_meals=fixed_meals,
        reroll_nonce=reroll_nonce,
        nutrition_entries=cache.nutrition_entries if cache is not None else None,
    )
    if rest and isinstance(rest.get("nutrients"), dict):
        meal_nutrients["午餐"] = {
            k: float(rest["nutrients"].get(k, 0.0)) for k in NUTRIENT_KEYS
        }
        # 餐廳午餐由餐廳固定營養值提供；唔應再用求解器午餐食材去計米類熟重。
        meal_items["午餐"] = []
        choice = str(rest.get("choice") or "").strip()
        store = str(rest.get("store") or "").strip()
        if choice or store:
            label = f'Lunch — "{choice}"'
            if store:
                label += f" ({store})"
            meal_ingredients["午餐"] = [label]
        else:
            meal_ingredients["午餐"] = ["Lunch — restaurant meal"]

    return {
        "primary_rule": primary_dict,
        "meal_patterns": by_meal,
        "meal_pattern_parts": meal_pattern_parts,
        "meal_ingredients": meal_ingredients,
        "meal_items": meal_items,
        "meal_nutrients": meal_nutrients,
        "rice_note": build_rice_note(meal_items, settings, visible_meals=visible_meals),
        "restaurant_lunch": rest,
        "meal_times_resolved": meal_times_resolved,
        "optimization": optimization_meta,
        "peninsula_stack_applied": False,
        "note": None,
    }
