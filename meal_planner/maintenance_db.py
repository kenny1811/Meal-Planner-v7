"""Generic SQLite maintenance copies for workbook sheets."""

from __future__ import annotations

from contextlib import closing
from datetime import date, datetime, time
import json
import sqlite3
from typing import Any

from openpyxl.workbook.workbook import Workbook

from meal_planner.excel_io import get_sheet
from meal_planner.settings import AppSettings, get_settings


MAINTENANCE_SHEETS: tuple[tuple[str, str], ...] = (
    ("roster", "更表"),
    ("overtime", "加班表"),
    ("payroll_times", "更時表"),
    ("public_holidays", "公眾假期"),
    ("medical_appointments", "醫療行程"),
    ("meal_times", "飯時表"),
    ("meal_patterns", "Pattern"),
    ("restaurant", "餐廳選擇"),
    ("schedule_grid", "行位表"),
)

RUNTIME_INPUT_SHEET_KEYS: tuple[str, ...] = ("roster", "overtime")


class MaintenanceDatabaseError(RuntimeError):
    """Maintenance data is unavailable and cannot be bootstrapped."""


def _connect(settings: AppSettings) -> sqlite3.Connection:
    path = settings.database_path
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS maintenance_sheets (
            sheet_key TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS maintenance_sheet_rows (
            sheet_key TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            cells_json TEXT NOT NULL,
            PRIMARY KEY (sheet_key, row_index),
            FOREIGN KEY (sheet_key) REFERENCES maintenance_sheets(sheet_key)
                ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS roster_code_definitions (
            pattern TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            sort_order INTEGER NOT NULL
        );
        """
    )
    conn.commit()


def _sheet_name_for_key(settings: AppSettings, sheet_key: str) -> str:
    mapping = {
        "roster": settings.sheets.roster,
        "overtime": settings.sheets.overtime,
        "payroll_times": settings.sheets.payroll_times,
        "public_holidays": settings.sheets.public_holidays,
        "medical_appointments": "醫療行程",
        "meal_times": settings.sheets.meal_times,
        "meal_patterns": settings.sheets.meal_times,
        "restaurant": settings.sheets.restaurant,
        "schedule_grid": settings.sheets.schedule_grid,
    }
    if sheet_key not in mapping:
        raise KeyError(sheet_key)
    return mapping[sheet_key]


def _display_name_for_key(sheet_key: str) -> str:
    for key, label in MAINTENANCE_SHEETS:
        if key == sheet_key:
            return label
    raise KeyError(sheet_key)


def _cell_to_json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return value


def _read_workbook_sheet_rows(settings: AppSettings, wb: Workbook, sheet_key: str) -> list[list[Any]]:
    sheet_name = _sheet_name_for_key(settings, sheet_key)
    ws = get_sheet(wb, sheet_name)
    max_row = int(ws.max_row or 0)
    
    if sheet_key == "meal_times":
        start_col = 1
        max_col = min(int(ws.max_column or 0), 5)
    elif sheet_key == "meal_patterns":
        start_col = 7
        max_col = min(int(ws.max_column or 0), 8)
    else:
        start_col = 1
        max_col = int(ws.max_column or 0)

    rows: list[list[Any]] = []
    for r in range(1, max_row + 1):
        row = [_cell_to_json_value(ws.cell(r, c).value) for c in range(start_col, max_col + 1)]
        while row and row[-1] is None:
            row.pop()
        rows.append(row)
    while rows and not any(cell not in (None, "") for cell in rows[-1]):
        rows.pop()
    return rows


def _has_sheet_rows(conn: sqlite3.Connection, sheet_key: str) -> bool:
    return (
        conn.execute(
            "SELECT 1 FROM maintenance_sheet_rows WHERE sheet_key = ? LIMIT 1",
            (sheet_key,),
        ).fetchone()
        is not None
    )


def save_sheet_rows(
    sheet_key: str,
    rows: list[list[Any]],
    settings: AppSettings | None = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    display_name = _display_name_for_key(sheet_key)
    source_sheet = _sheet_name_for_key(settings, sheet_key)
    clean_rows = [list(row) if isinstance(row, list) else [] for row in rows]
    while clean_rows and not any(cell not in (None, "") for cell in clean_rows[-1]):
        clean_rows.pop()

    now = datetime.now().isoformat(timespec="seconds")
    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        conn.execute(
            """
            INSERT OR REPLACE INTO maintenance_sheets(
                sheet_key, display_name, source_sheet, updated_at
            ) VALUES (?, ?, ?, ?)
            """,
            (sheet_key, display_name, source_sheet, now),
        )
        conn.execute("DELETE FROM maintenance_sheet_rows WHERE sheet_key = ?", (sheet_key,))
        conn.executemany(
            """
            INSERT INTO maintenance_sheet_rows(sheet_key, row_index, cells_json)
            VALUES (?, ?, ?)
            """,
            [
                (sheet_key, idx, json.dumps(row, ensure_ascii=False))
                for idx, row in enumerate(clean_rows, start=1)
            ],
        )
        conn.commit()
    return {"sheet_key": sheet_key, "updated_at": now, "row_count": len(clean_rows)}


def bootstrap_sheet_from_workbook(
    settings: AppSettings,
    wb: Workbook,
    sheet_key: str,
    *,
    replace_existing: bool = False,
) -> None:
    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        if not replace_existing and _has_sheet_rows(conn, sheet_key):
            return
    save_sheet_rows(sheet_key, _read_workbook_sheet_rows(settings, wb, sheet_key), settings)


def bootstrap_all_from_workbook(
    settings: AppSettings,
    wb: Workbook,
    *,
    replace_existing: bool = False,
) -> None:
    for sheet_key, _ in MAINTENANCE_SHEETS:
        bootstrap_sheet_from_workbook(settings, wb, sheet_key, replace_existing=replace_existing)
    bootstrap_roster_code_definitions(settings, wb)


def import_runtime_inputs_from_workbook(
    settings: AppSettings,
    wb: Workbook,
    *,
    replace_existing: bool = True,
) -> dict[str, Any]:
    """Import live operational inputs used by preview generation."""
    imported: list[dict[str, Any]] = []
    for sheet_key in RUNTIME_INPUT_SHEET_KEYS:
        bootstrap_sheet_from_workbook(settings, wb, sheet_key, replace_existing=replace_existing)
        imported.append(load_sheet_rows(sheet_key, settings))
    bootstrap_roster_code_definitions(settings, wb, replace_existing=replace_existing)
    return {
        "runtime_input_keys": list(RUNTIME_INPUT_SHEET_KEYS),
        "sheets": [
            {
                "sheet_key": item["sheet_key"],
                "display_name": item["display_name"],
                "source_sheet": item["source_sheet"],
                "updated_at": item["updated_at"],
                "row_count": len(item.get("rows", [])),
            }
            for item in imported
        ],
        "roster_code_definitions": load_roster_code_definitions(settings),
    }


def list_runtime_input_status(settings: AppSettings | None = None) -> dict[str, Any]:
    """Return import status for the SQLite-backed live input sheets."""
    settings = settings or get_settings()
    all_sheets = {
        str(sheet["sheet_key"]): sheet
        for sheet in list_maintenance_sheets(settings)
    }
    return {
        "excel_role": "import_only",
        "runtime_input_keys": list(RUNTIME_INPUT_SHEET_KEYS),
        "sheets": [all_sheets[key] for key in RUNTIME_INPUT_SHEET_KEYS],
    }


def _read_roster_code_definitions_from_workbook(settings: AppSettings, wb: Workbook) -> list[dict[str, Any]]:
    ws = get_sheet(wb, settings.sheets.roster)
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for r in range(1, int(ws.max_row or 0) + 1):
        pattern = str(ws.cell(r, 3).value or "").strip()
        label = str(ws.cell(r, 4).value or "").strip()
        if not pattern or not label or pattern == "更碼":
            continue
        if pattern in seen:
            continue
        seen.add(pattern)
        rows.append({"pattern": pattern, "label": label, "sort_order": len(rows) + 1})
    return rows


def _has_roster_code_definitions(conn: sqlite3.Connection) -> bool:
    return conn.execute("SELECT 1 FROM roster_code_definitions LIMIT 1").fetchone() is not None


def bootstrap_roster_code_definitions(
    settings: AppSettings,
    wb: Workbook,
    *,
    replace_existing: bool = False,
) -> None:
    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        if not replace_existing and _has_roster_code_definitions(conn):
            return
    save_roster_code_definitions(_read_roster_code_definitions_from_workbook(settings, wb), settings)


def load_roster_code_definitions(settings: AppSettings | None = None) -> list[dict[str, Any]]:
    settings = settings or get_settings()
    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        rows = conn.execute(
            """
            SELECT pattern, label, sort_order
            FROM roster_code_definitions
            ORDER BY sort_order, pattern
            """
        ).fetchall()
    return [
        {
            "pattern": str(row["pattern"]),
            "label": str(row["label"]),
            "sort_order": int(row["sort_order"]),
        }
        for row in rows
    ]


def save_roster_code_definitions(
    rows: list[dict[str, Any]],
    settings: AppSettings | None = None,
) -> list[dict[str, Any]]:
    settings = settings or get_settings()
    clean_rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for idx, raw in enumerate(rows, start=1):
        if not isinstance(raw, dict):
            continue
        pattern = str(raw.get("pattern") or "").strip()
        label = str(raw.get("label") or "").strip()
        if not pattern and not label:
            continue
        if not pattern or not label:
            raise ValueError(f"Roster code definition row {idx} requires Pattern and Definition.")
        if pattern in seen:
            raise ValueError(f"Roster code definition pattern is duplicated: {pattern}")
        seen.add(pattern)
        clean_rows.append({"pattern": pattern, "label": label, "sort_order": len(clean_rows) + 1})

    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        conn.execute("DELETE FROM roster_code_definitions")
        conn.executemany(
            """
            INSERT INTO roster_code_definitions(pattern, label, sort_order)
            VALUES (:pattern, :label, :sort_order)
            """,
            clean_rows,
        )
        conn.commit()
    return load_roster_code_definitions(settings)


def list_maintenance_sheets(settings: AppSettings | None = None) -> list[dict[str, Any]]:
    settings = settings or get_settings()
    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        meta = {
            str(row["sheet_key"]): row
            for row in conn.execute("SELECT * FROM maintenance_sheets").fetchall()
        }
        counts = {
            str(row["sheet_key"]): int(row["row_count"])
            for row in conn.execute(
                """
                SELECT sheet_key, COUNT(*) AS row_count
                FROM maintenance_sheet_rows
                GROUP BY sheet_key
                """
            ).fetchall()
        }
    out: list[dict[str, Any]] = []
    for sheet_key, display_name in MAINTENANCE_SHEETS:
        row = meta.get(sheet_key)
        out.append(
            {
                "sheet_key": sheet_key,
                "display_name": display_name,
                "source_sheet": _sheet_name_for_key(settings, sheet_key),
                "updated_at": row["updated_at"] if row else None,
                "row_count": counts.get(sheet_key, 0),
            }
        )
    return out


def load_sheet_rows(
    sheet_key: str,
    settings: AppSettings | None = None,
    wb: Workbook | None = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        has_rows = _has_sheet_rows(conn, sheet_key)
    if not has_rows:
        if wb is None:
            raise MaintenanceDatabaseError(
                f"Maintenance sheet {_display_name_for_key(sheet_key)} is empty and no workbook was provided."
            )
        bootstrap_sheet_from_workbook(settings, wb, sheet_key)

    with closing(_connect(settings)) as conn:
        _ensure_schema(conn)
        meta = conn.execute(
            "SELECT * FROM maintenance_sheets WHERE sheet_key = ?",
            (sheet_key,),
        ).fetchone()
        rows = conn.execute(
            """
            SELECT row_index, cells_json
            FROM maintenance_sheet_rows
            WHERE sheet_key = ?
            ORDER BY row_index
            """,
            (sheet_key,),
        ).fetchall()

    return {
        "sheet_key": sheet_key,
        "display_name": _display_name_for_key(sheet_key),
        "source_sheet": _sheet_name_for_key(settings, sheet_key),
        "updated_at": meta["updated_at"] if meta else None,
        "rows": [json.loads(row["cells_json"]) for row in rows],
    }
